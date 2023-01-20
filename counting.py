import glob
import os
import openpyxl as xls

if __name__ == '__main__':
    # folder_list = glob.glob(os.getcwd() + '/data/api_links/*')
    # folder_list = glob.glob(os.getcwd() + '\\extract_data\\*')
    folder_list = glob.glob(os.getcwd() + '/extract_data/*')
    api_list = glob.glob(os.getcwd() + '/data/api_links/*')

    wb = xls.Workbook()
    ws = wb.active
    ws['A1'] = '회사명'
    ws['B1'] = '갯수'
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 50

    total_count = 0
    row_count = 2
    for row_idx, folder_path in enumerate(folder_list):
        company = folder_path.split("/")[-1]
        txt_file_list = glob.glob(folder_path + '/*.txt')
        counting = 0
        for txt_file_path in txt_file_list:
            f = open(txt_file_path, 'r', encoding='utf-8')

            counting = counting + len(f.readlines())
        row_count = row_count + 1
        print(row_count)
        ws['A{}'.format(row_count)] = company
        ws['B{}'.format(row_count)] = counting

        total_count += counting

    extract_list = list(map(lambda s: s.split('/')[-1], folder_list))
    all_list = list(map(lambda s: s.split('/')[-1], api_list))


    temp3 = []
    for i in all_list:
        if '大阪' in i:
            continue
        if '東京' in i:
            continue
        if i not in extract_list:
            temp3.append(i)
    for company in temp3:
        ws['A{}'.format(row_count + 1)] = company
        ws['B{}'.format(row_count + 1)] = 0
        row_count += 1

    ws['A{}'.format(row_count + 2)] = '총갯수'
    ws['B{}'.format(row_count + 2)] = total_count

    wb.save('./ttt.xlsx')
    wb.close()
