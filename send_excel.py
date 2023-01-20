import glob
import json
import os
import openpyxl as xls
import re
from tqdm import tqdm


def cleanhtml(raw_html):
    cleanr = re.compile('<.*?>')
    cleantext = re.sub(cleanr, '', raw_html)
    return cleantext


if __name__ == '__main__':

    company_folder_list = glob.glob('{}/{}/*'.format(os.getcwd(), 'extract_data'))
    xlsx_files = glob.glob('./xlsx_files/*')
    xlsx_files = list(
        map(lambda s: s.replace('./xlsx_files/', '').replace('.xlsx', ''), xlsx_files))

    for company_index, company_folder in enumerate(company_folder_list):

        company_name = company_folder.split('/')[-1]

        if company_name in xlsx_files:
            continue

        wb = xls.Workbook()

        country_txt_file_path_list = glob.glob('{}/{}/{}/*.txt'.format(
            os.getcwd(), 'extract_data', company_name
        ))

        for i, file_path in enumerate(tqdm(country_txt_file_path_list)):
            country_name = file_path.split('/')[-1].replace('.txt', '').replace('\n', '')

            print('start : ', company_name, country_name)
            ws = wb.create_sheet(index=i, title=country_name)
            ws['A1'] = 'gubun'
            ws['B1'] = 'companyName'
            ws['C1'] = 'region'
            ws['D1'] = 'Job_Title'
            ws['E1'] = 'Salary'
            ws['F1'] = 'Full Job Description'
            ws['G1'] = 'url'
            ws.column_dimensions["A"].width = 50
            ws.column_dimensions["B"].width = 50
            ws.column_dimensions["C"].width = 50
            ws.column_dimensions["D"].width = 50
            ws.column_dimensions["E"].width = 50
            ws.column_dimensions["F"].width = 50
            ws.column_dimensions["G"].width = 50

            country_file = open(file_path, 'r', encoding='utf-8')

            json_list = list(map(lambda s: s.strip(), country_file.readlines()))

            for j, data in enumerate(tqdm(json_list)):
                data_dict = json.loads(data)
                body = data_dict['body']
                job_info_model = body['jobInfoWrapperModel']['jobInfoModel']

                gubun = 'indeed'
                company_name = company_name
                region = body['jobLocation']
                job_title = job_info_model['jobInfoHeaderModel']['jobTitle']
                url = 'https://{}.indeed.com/cmp/{}/jobs?jk={}'.format(country_name, company_name,
                                                                       body['jobKey'])
                salary = None
                s_currency = None
                s_max = None
                s_min = None
                s_type = None
                try:
                    job_info_model = body['jobInfoWrapperModel']['jobInfoModel']
                    s_currency = job_info_model['jobInfoHeaderModel']['salaryCurrency']
                    s_max = job_info_model['jobInfoHeaderModel']['salaryMax']
                    s_min = job_info_model['jobInfoHeaderModel']['salaryMin']
                    s_type = job_info_model['jobInfoHeaderModel']['salaryType']
                    salary = "[{}]\nMin:{}\nMax:{}\ntype:{}".format(s_currency, s_min,
                                                                    s_max, s_type)
                except Exception as e:
                    pass

                if s_max is None and s_min is None:
                    try:
                        sgm_range = body['salaryGuideModel']['estimatedSalaryModel']['formattedRange']
                        sgm_max = body['salaryGuideModel']['estimatedSalaryModel']['max']
                        sgm_min = body['salaryGuideModel']['estimatedSalaryModel']['min']
                        sgm_type = body['salaryGuideModel']['estimatedSalaryModel']['type']

                        if sgm_max is None and sgm_min is None:
                            salary = None
                        else:
                            salary = "[{}]\nMin:{}\nMax:{}\ntype:{}".format(sgm_range, sgm_min,
                                                                            sgm_max, sgm_type)
                    except:
                        salary = None

                try:
                    job_type = \
                        job_info_model['jobDescriptionSectionModel']['jobDetailsSection'][
                            'contents']['Job Type']

                    if job_type is None:
                        job_list = \
                        job_info_model['jobDescriptionSectionModel']['jobDetailsSection'][
                            'jobTypes']
                        jobs = []
                        for job_item in job_list:
                            jobs.append(job_item['label'])
                        job_type = jobs

                    job_type = '\n'.join(job_type)
                except:
                    job_type = None

                description = cleanhtml(job_info_model['sanitizedJobDescription']['content'])

                ws['A{}'.format(j + 2)] = gubun
                ws['B{}'.format(j + 2)] = company_name
                ws['C{}'.format(j + 2)] = region
                ws['D{}'.format(j + 2)] = job_title
                ws['E{}'.format(j + 2)] = salary
                ws['F{}'.format(j + 2)] = description
                ws['G{}'.format(j + 2)] = url

                wb.save('./xlsx_files/{}.xlsx'.format(company_name))
        wb.close()
